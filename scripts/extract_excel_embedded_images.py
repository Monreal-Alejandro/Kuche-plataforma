#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrae imágenes incrustadas de un .xlsx y las guarda nombradas según el texto
de una columna en la misma fila que el ancla de la imagen.

Uso típico (desde la raíz del proyecto):
  py scripts/extract_excel_embedded_images.py ELECTRODOMESTICOS.xlsx

Requisitos:
  py -m pip install -r scripts/requirements-excel-images.txt

Si en Windows `py` apunta a una ruta que ya no existe o `python` abre la Tienda:
  - Reinstala Python desde https://www.python.org/downloads/ (marca "Add to PATH"), o
  - Usa la alternativa sin Python: npm run extract:excel-images (ver scripts/extract-excel-embedded-images.mjs).
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
import zipfile
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Falta openpyxl. Instala con: py -m pip install -r scripts/requirements-excel-images.txt", file=sys.stderr)
    sys.exit(1)


def sanitize_stem(text: object) -> str:
    """Minúsculas, espacios -> guiones bajos, quita caracteres no seguros para nombre de archivo."""
    if text is None or str(text).strip() == "":
        return "sin_nombre"
    t = unicodedata.normalize("NFKD", str(text))
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.lower().strip()
    t = re.sub(r"\s+", "_", t)
    t = re.sub(r'[<>:"/\\|?*]+', "", t)
    t = re.sub(r"[^a-z0-9_.-]+", "_", t)
    t = re.sub(r"_+", "_", t).strip("_")
    if not t:
        return "sin_nombre"
    return t[:180]


def anchor_to_row_col(anchor) -> tuple[int | None, int | None]:
    """Devuelve fila y columna 1-based del ancla superior izquierda de la imagen."""
    if anchor is None:
        return None, None
    fr = getattr(anchor, "_from", None)
    if fr is None:
        return None, None
    try:
        return int(fr.row) + 1, int(fr.col) + 1
    except (TypeError, ValueError):
        return None, None


def zip_member_path(path_attr: str) -> str:
    """Normaliza ruta tipo '/xl/media/image1.png' -> 'xl/media/image1.png' para ZipFile."""
    if not path_attr:
        return ""
    p = path_attr.replace("\\", "/").strip()
    if p.startswith("/"):
        p = p[1:]
    return p


def read_image_bytes(xlsx_path: Path, img) -> tuple[bytes, str]:
    """
    Obtiene bytes y extensión sugerida (.png, .jpeg, …).
    openpyxl guarda la ruta del recurso en el paquete ZIP del xlsx.
    """
    # API interna en algunas versiones
    if hasattr(img, "_data"):
        try:
            data = img._data()
            if isinstance(data, bytes) and data:
                return data, ".png"
        except Exception:
            pass

    # Algunas versiones exponen bytes o buffer en .ref
    ref = getattr(img, "ref", None)
    if isinstance(ref, (bytes, bytearray)):
        return bytes(ref), ".png"
    if ref is not None and hasattr(ref, "read"):
        try:
            data = ref.read()
            if isinstance(data, bytes):
                return data, ".png"
        except Exception:
            pass

    path_attr = getattr(img, "path", None)
    if isinstance(ref, str) and ref:
        path_attr = ref
    if not path_attr or not isinstance(path_attr, str):
        raise ValueError(
            "No se encontró ruta de imagen en el objeto (¿versión de openpyxl?). "
            f"Tipo: {type(img).__name__}. Atributos: path={getattr(img, 'path', None)!r}"
        )

    member = zip_member_path(path_attr)
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        names = zf.namelist()
        if member not in names:
            # A veces la ruta viene sin prefijo xl/
            base = member.split("/")[-1]
            candidates = [n for n in names if n.endswith(base) and "media" in n.replace("\\", "/")]
            if len(candidates) == 1:
                member = candidates[0]
            elif member in [n.replace("\\", "/") for n in names]:
                pass
            else:
                raise FileNotFoundError(f"No está en el ZIP del xlsx: {member!r} (¿imagen enlazada externa?)")

        data = zf.read(member)
        ext = Path(member).suffix.lower()
        if ext not in {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tif", ".tiff", ".emf", ".wmf"}:
            ext = ".bin"
        return data, ext


def unique_path(folder: Path, stem: str, ext: str) -> Path:
    base = folder / f"{stem}{ext}"
    if not base.exists():
        return base
    n = 2
    while True:
        cand = folder / f"{stem}_{n}{ext}"
        if not cand.exists():
            return cand
        n += 1


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extrae imágenes incrustadas de Excel y las nombra según una columna de la misma fila."
    )
    parser.add_argument(
        "excel",
        nargs="?",
        default="ELECTRODOMESTICOS.xlsx",
        help="Ruta al .xlsx (por defecto: ELECTRODOMESTICOS.xlsx en el directorio actual)",
    )
    parser.add_argument(
        "-o",
        "--out",
        default="imagenes_extraidas",
        help="Carpeta de salida (por defecto: imagenes_extraidas)",
    )
    parser.add_argument(
        "-c",
        "--column",
        type=int,
        default=1,
        metavar="N",
        help="Columna 1-based del nombre del material (1=A, 2=B, …). Por defecto: 1",
    )
    parser.add_argument(
        "-s",
        "--sheet",
        default=None,
        help="Nombre de la hoja. Si se omite, usa la hoja activa del libro.",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.excel).resolve()
    if not xlsx_path.is_file():
        print(f"No existe el archivo: {xlsx_path}", file=sys.stderr)
        return 1

    out_dir = Path(args.out).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, read_only=False, keep_links=True, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active

    images = getattr(ws, "_images", None)
    if not images:
        print(
            "No se encontraron imágenes incrustadas en esta hoja (_images vacío).\n"
            "Comprueba: (1) que las fotos estén insertadas en Excel como imagen en la hoja, "
            "(2) que sea la hoja correcta (-s), (3) que el archivo no sea solo .xls antiguo.",
            file=sys.stderr,
        )
        return 2

    saved = 0
    errors = 0
    for idx, img in enumerate(images):
        try:
            row, _col = anchor_to_row_col(getattr(img, "anchor", None))
            if row is None:
                print(f"[{idx + 1}] Sin ancla de fila, se omite.", file=sys.stderr)
                errors += 1
                continue

            cell_val = ws.cell(row=row, column=args.column).value
            stem = sanitize_stem(cell_val)
            data, ext = read_image_bytes(xlsx_path, img)
            dest = unique_path(out_dir, stem, ext)
            dest.write_bytes(data)
            print(f"OK  fila {row}  ->  {dest.name}")
            saved += 1
        except Exception as e:
            print(f"[{idx + 1}] Error: {e}", file=sys.stderr)
            errors += 1

    print(f"\nListo: {saved} imagen(es) en {out_dir}")
    if errors:
        print(f"Avisos/errores: {errors}", file=sys.stderr)
    return 0 if saved else 3


if __name__ == "__main__":
    raise SystemExit(main())
